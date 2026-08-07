"""Build 400-event restoration phase Excel workbook with raw-log enrichment."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from restoration_raw_lookup import (
    build_session_event_index,
    lookup_raw_event,
    resolve_jsonl_path,
)


ROOT = Path(__file__).resolve().parents[1]
THIRDPASS_DIR = ROOT / "analysis_data" / "model_runs" / "thirdpass_model_benchmark_benign1to4"
OUT_XLSX = ROOT / "docs_active" / "復元フェーズ_400ログ台帳.xlsx"

SCENARIOS = ("s3", "m4", "m6", "s4")
COARSE_CHUNK_SIZE = 100
MICRO_CHUNK_SIZE = 10

FILL_ATTACK = PatternFill("solid", fgColor="FFC7CE")
FILL_UC = PatternFill("solid", fgColor="C6EFCE")
FILL_CAUTION = PatternFill("solid", fgColor="FFEB9C")
FILL_AUX = PatternFill("solid", fgColor="FFF2CC")
FILL_EXCLUDE = PatternFill("solid", fgColor="D9D9D9")
FILL_CONTEXT = PatternFill("solid", fgColor="E7E6E6")
FILL_HEADER = PatternFill("solid", fgColor="4472C4")
FONT_HEADER = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

BACKGROUND_PROCESSES = {"sysmon.exe", "csrss.exe", "svchost.exe", "services.exe"}
CONTEXT_PROCESSES = {"vmtoolsd.exe", "searchindexer.exe"}
HOLD_PROCESSES = {"searchfilterhost.exe", "wmiprvse.exe"}

EVENT_ID_JA = {
    "4656": "ハンドル要求",
    "4663": "オブジェクトアクセス",
    "4658": "ハンドル解放",
    "4660": "オブジェクト削除",
    "4688": "プロセス作成(監査)",
    "4690": "プロセス作成(詳細)",
}

PROCESS_META: dict[str, tuple[str, str]] = {
    "repmgr.exe": ("レプリケーション管理", "管理エージェント"),
    "repwmiutils.exe": ("WMI収集ユーティリティ", "管理・収集"),
    "tpautoconnsvc.exe": ("自動接続サービス", "名前から役割不明→要確認"),
    "searchprotocolhost.exe": ("Windows検索プロトコル", "索引・検索"),
    "searchfilterhost.exe": ("検索フィルタ", "索引補助"),
    "searchindexer.exe": ("検索インデクサ", "索引バックグラウンド"),
    "spoolsv.exe": ("スプーラー", "印刷スプール"),
    "sysmon.exe": ("Sysmon", "背景・監視"),
    "csrss.exe": ("CSRSS", "背景OS"),
    "svchost.exe": ("サービスホスト", "背景OS"),
    "services.exe": ("サービス制御", "背景OS"),
    "vmtoolsd.exe": ("VMware Tools", "仮想環境"),
    "wmiprvse.exe": ("WMIプロバイダ", "WMI"),
    "": ("(プロセス名なし)", "監査のみ"),
}

RAW_KEYS = [
    "@timestamp",
    "EventRecordID",
    "EventID",
    "Channel",
    "Computer",
    "ProcessName",
    "ObjectName",
    "ObjectType",
    "AccessMask",
    "AccessList",
    "HandleId",
    "SubjectUserName",
    "SubjectDomainName",
    "SubjectLogonId",
    "ProcessId",
    "ThreadID",
]

USECASE_ROWS = [
    {
        "usecase_id": "U1",
        "pattern": "repmgr.exe / EventID=4663 / ObjectType=file",
        "process": "repmgr.exe",
        "event_id": "4663",
        "object_type": "file",
        "priority_scenarios": "M4, S4",
        "restore_status": "未着手",
        "representative_key": "",
        "representative_object_name": "",
        "notes": "最頻出・管理エージェント file access",
    },
    {
        "usecase_id": "U2",
        "pattern": "repwmiutils.exe / EventID=4663 / ObjectType=file",
        "process": "repwmiutils.exe",
        "event_id": "4663",
        "object_type": "file",
        "priority_scenarios": "M4, M6",
        "restore_status": "未着手",
        "representative_key": "",
        "representative_object_name": "",
        "notes": "収集補助・U1とセットで管理系帯",
    },
    {
        "usecase_id": "U3",
        "pattern": "tpautoconnsvc.exe / EventID=4663 / ObjectType=file",
        "process": "tpautoconnsvc.exe",
        "event_id": "4663",
        "object_type": "file",
        "priority_scenarios": "S4, S3",
        "restore_status": "未着手",
        "representative_key": "",
        "representative_object_name": "",
        "notes": "単独は背景寄り。winword/repmgr 同居 micro を選ぶ",
    },
    {
        "usecase_id": "U4",
        "pattern": "searchprotocolhost.exe / EventID=4663 / ObjectType=file",
        "process": "searchprotocolhost.exe",
        "event_id": "4663",
        "object_type": "file",
        "priority_scenarios": "M4, S4",
        "restore_status": "未着手",
        "representative_key": "",
        "representative_object_name": "",
        "notes": "selected with caution / 境界正常",
    },
    {
        "usecase_id": "補助",
        "pattern": "spoolsv.exe / EventID=4660 or 4663 / ObjectType=file",
        "process": "spoolsv.exe",
        "event_id": "4663",
        "object_type": "file",
        "priority_scenarios": "M4",
        "restore_status": "未着手",
        "representative_key": "",
        "representative_object_name": "",
        "notes": "主採用より一段下",
    },
]

ATTACK_ROWS = [
    {
        "scenario": "M4",
        "micro_rank": 3,
        "parent_session_id": "win-32-h1|win-32-h1$|20220719T2250Z",
        "coarse_chunk_index": 68,
        "micro_chunk_index": 5,
        "offset_in_micro": 4,
        "template": "4663 | searchprotocolhost.exe | file",
        "groundtruth_checked": "",
        "object_name": "",
        "sysmon_note": "",
        "restore_layer_a": "未",
        "restore_layer_b": "未",
        "notes": "同 micro: searchfilterhost, repmgr 等 normal 9件",
        "raw_one_liner": "",
    },
    {
        "scenario": "S4",
        "micro_rank": 9,
        "parent_session_id": "win-32-h1|win-32-h1$|20220720T0050Z",
        "coarse_chunk_index": 10,
        "micro_chunk_index": 9,
        "offset_in_micro": 6,
        "template": "4663 | searchprotocolhost.exe | file",
        "groundtruth_checked": "",
        "object_name": "",
        "sysmon_note": "",
        "restore_layer_a": "未",
        "restore_layer_b": "未",
        "notes": "同 micro: sysmon, wmiprvse 等 normal 9件",
        "raw_one_liner": "",
    },
]

CHECKLIST_ROWS = [
    ["#", "対象", "調査項目", "データソース", "完了"],
    ["1", "全行", "採用区分・プロセス表示列で読む", "01_400マスター", ""],
    ["2", "micro", "10件並びを塊で理解", "01b_micro10一覧", ""],
    ["3", "U1-U4", "ObjectName・ユーザー・AccessMask", "生ログJSON列", ""],
    ["4", "U1-U4", "復元成立判定（方針§4）", "docs_active/復元フェーズ方針", ""],
    ["5", "残attack2", "groundtruth 照合", "04_残attack2", ""],
]

READABLE_HEADERS = [
    "row_id",
    "scenario",
    "adoption",
    "label_name",
    "process_display",
    "process_category",
    "EventID",
    "event_action_ja",
    "ObjectType",
    "ObjectName",
    "@timestamp",
    "SubjectUserName",
    "AccessMask",
    "EventRecordID",
    "raw_one_liner",
    "micro_rank",
    "offset_in_micro",
    "micro_10_summary",
    "parent_session_id",
    "minute_bucket",
    "raw_lookup_status",
    "raw_json",
    "template",
    "restore_status",
    "event_key",
    "offset_in_session",
    "coarse_chunk_index",
    "micro_chunk_index",
    "HandleId",
    "Computer",
    "color_hint",
]


def parse_template(template: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for segment in template.split("|"):
        segment = segment.strip()
        if "=" in segment:
            key, value = segment.split("=", 1)
            out[key.strip()] = value.strip()
    return out


def process_display_name(process: str) -> tuple[str, str]:
    key = process or ""
    label, category = PROCESS_META.get(key, (key or "(不明)", "未分類"))
    if key and key not in PROCESS_META:
        return f"{key}（要調査）", "未分類"
    if key:
        return f"{key}（{label}）", category
    return PROCESS_META[""][0], PROCESS_META[""][1]


def classify_row(process: str, event_id: str, object_type: str, label: int) -> tuple[str, str]:
    if label == 1:
        return "attack", "赤"
    if not process and event_id in {"4688", "4690"}:
        return "除外", "灰"
    if process in BACKGROUND_PROCESSES:
        return "除外", "灰"
    if process == "searchprotocolhost.exe" and event_id == "4663" and object_type == "file":
        return "U4", "黄"
    if process == "repmgr.exe" and event_id == "4663" and object_type == "file":
        return "U1", "緑"
    if process == "repwmiutils.exe" and event_id == "4663" and object_type == "file":
        return "U2", "緑"
    if process == "tpautoconnsvc.exe" and event_id == "4663" and object_type == "file":
        return "U3", "緑"
    if process == "spoolsv.exe" and event_id in {"4660", "4663"}:
        return "補助", "黄"
    if process in HOLD_PROCESSES:
        return "保留", "白"
    if process in CONTEXT_PROCESSES:
        return "補助文脈", "灰"
    uc_near = {
        "repmgr.exe": "U1_周辺",
        "repwmiutils.exe": "U2_周辺",
        "tpautoconnsvc.exe": "U3_周辺",
        "searchprotocolhost.exe": "U4_周辺",
        "spoolsv.exe": "補助_周辺",
    }
    if process in uc_near:
        return uc_near[process], "白"
    return "未分類", "白"


def fill_for_adoption(adoption: str, color_hint: str) -> PatternFill | None:
    if adoption == "attack":
        return FILL_ATTACK
    if adoption in {"U1", "U2", "U3"}:
        return FILL_UC
    if adoption == "U4":
        return FILL_CAUTION
    if adoption in {"補助", "保留"}:
        return FILL_AUX
    if adoption in {"除外", "補助文脈"}:
        return FILL_EXCLUDE
    if color_hint == "灰":
        return FILL_CONTEXT
    return None


def format_event_line(row: dict[str, object], include_object: bool = True) -> str:
    tag = "攻撃" if row.get("label") == 1 else "正常"
    proc = row.get("ProcessName") or "(なし)"
    eid = row.get("EventID", "")
    eja = row.get("event_action_ja") or EVENT_ID_JA.get(str(eid), "")
    obj = row.get("ObjectName") or row.get("ObjectType") or ""
    if include_object and obj:
        tail = f" → {obj}"
    else:
        tail = f" [{row.get('ObjectType', '')}]" if row.get("ObjectType") else ""
    return f"[{tag}] {proc} | {eid} {eja}{tail}"


def build_raw_one_liner(row: dict[str, object]) -> str:
    parts = [
        str(row.get("@timestamp") or row.get("minute_bucket") or "?"),
        "攻撃" if row.get("label") == 1 else "正常",
        str(row.get("process_display") or row.get("ProcessName") or "(なし)"),
        f"EID{row.get('EventID')} {row.get('event_action_ja', '')}".strip(),
    ]
    if row.get("ObjectName"):
        parts.append(f"Object={row['ObjectName']}")
    elif row.get("ObjectType"):
        parts.append(f"type={row['ObjectType']}")
    if row.get("SubjectUserName"):
        parts.append(f"User={row['SubjectUserName']}")
    if row.get("AccessMask"):
        parts.append(f"Access={row['AccessMask']}")
    return " | ".join(parts)


def compact_raw_json(event: dict | None) -> str:
    if not event:
        return ""
    payload = {k: event.get(k, "") for k in RAW_KEYS if event.get(k, "") not in (None, "")}
    return json.dumps(payload, ensure_ascii=False)


def enrich_from_raw(row: dict[str, object], event: dict | None, lookup_status: str) -> None:
    row["raw_lookup_status"] = lookup_status
    if not event:
        return
    row["@timestamp"] = event.get("@timestamp", "")
    row["ObjectName"] = event.get("ObjectName", "") or event.get("TargetFilename", "")
    row["SubjectUserName"] = event.get("SubjectUserName", "")
    row["AccessMask"] = event.get("AccessMask", "") or event.get("AccessList", "")
    row["EventRecordID"] = event.get("EventRecordID", "")
    row["HandleId"] = event.get("HandleId", "")
    row["Computer"] = event.get("Computer", "")
    row["raw_json"] = compact_raw_json(event)
    if event.get("ProcessName"):
        row["ProcessName"] = event["ProcessName"].split("\\")[-1].lower()
    row["raw_one_liner"] = build_raw_one_liner(row)


def global_offset(coarse_chunk_index: int, micro_chunk_index: int, offset_in_micro: int) -> int:
    return (
        coarse_chunk_index * COARSE_CHUNK_SIZE
        + micro_chunk_index * MICRO_CHUNK_SIZE
        + (offset_in_micro - 1)
    )


def load_master_rows(
    session_indexes: dict[str, dict[str, list[dict]]],
    lookup_notes: dict[str, str],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    event_idx = 0
    for scenario in SCENARIOS:
        csv_path = THIRDPASS_DIR / f"{scenario}_knn_review100" / "top_micro_flat_events.csv"
        if not csv_path.exists():
            raise FileNotFoundError(csv_path)
        session_index = session_indexes.get(scenario, {})
        lookup_status = lookup_notes.get(scenario, "JSONL未読込")

        with csv_path.open(encoding="utf-8-sig", newline="") as f:
            for raw in csv.DictReader(f):
                event_idx += 1
                fields = parse_template(raw["template"])
                process = fields.get("ProcessName", "")
                event_id = fields.get("EventID", "")
                object_type = fields.get("ObjectType", "")
                label = int(raw["label"])
                adoption, color_hint = classify_row(process, event_id, object_type, label)
                disp, category = process_display_name(process)
                micro_key = (
                    f"{scenario}|{raw['parent_session_id']}|c{raw['coarse_chunk_index']}"
                    f"|m{raw['micro_chunk_index']}"
                )
                event_key = f"{micro_key}|o{raw['offset_in_micro']}"
                sess = raw["parent_session_id"]
                minute = sess.split("|")[-1] if "|" in sess else sess
                offset_sess = global_offset(
                    int(raw["coarse_chunk_index"]),
                    int(raw["micro_chunk_index"]),
                    int(raw["offset_in_micro"]),
                )

                row: dict[str, object] = {
                    "row_id": event_idx,
                    "scenario": scenario.upper(),
                    "micro_rank": int(raw["rank"]),
                    "offset_in_micro": int(raw["offset_in_micro"]),
                    "offset_in_session": offset_sess,
                    "label": label,
                    "label_name": "attack" if label == 1 else "normal",
                    "template": raw["template"],
                    "ProcessName": process,
                    "process_display": disp,
                    "process_category": category,
                    "EventID": event_id,
                    "event_action_ja": EVENT_ID_JA.get(event_id, ""),
                    "ObjectType": object_type,
                    "parent_session_id": sess,
                    "minute_bucket": minute,
                    "coarse_chunk_rank": int(raw["coarse_chunk_rank"]),
                    "coarse_chunk_index": int(raw["coarse_chunk_index"]),
                    "micro_chunk_index": int(raw["micro_chunk_index"]),
                    "pass_stage": "third_100",
                    "adoption": adoption,
                    "color_hint": color_hint,
                    "restore_status": "未着手",
                    "micro_chunk_key": micro_key,
                    "event_key": event_key,
                    "ObjectName": "",
                    "@timestamp": "",
                    "SubjectUserName": "",
                    "AccessMask": "",
                    "EventRecordID": "",
                    "HandleId": "",
                    "Computer": "",
                    "raw_one_liner": build_raw_one_liner(
                        {
                            "label": label,
                            "process_display": disp,
                            "ProcessName": process,
                            "EventID": event_id,
                            "event_action_ja": EVENT_ID_JA.get(event_id, ""),
                            "ObjectType": object_type,
                            "minute_bucket": minute,
                        }
                    ),
                    "raw_json": "",
                    "micro_10_summary": "",
                    "raw_lookup_status": lookup_status,
                    "notes": "",
                }

                raw_event = lookup_raw_event(
                    session_index,
                    sess,
                    int(raw["coarse_chunk_index"]),
                    int(raw["micro_chunk_index"]),
                    int(raw["offset_in_micro"]),
                )
                enrich_from_raw(row, raw_event, lookup_status if raw_event else f"{lookup_status} (index miss)")
                row["raw_one_liner"] = build_raw_one_liner(row)
                rows.append(row)
    return rows


def attach_micro_summaries(rows: list[dict[str, object]]) -> None:
    groups: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        groups[str(row["micro_chunk_key"])].append(row)
    for group in groups.values():
        group.sort(key=lambda r: int(r["offset_in_micro"]))
        lines = []
        for item in group:
            lines.append(f"{int(item['offset_in_micro']):02d}. {format_event_line(item)}")
        summary = "\n".join(lines)
        proc_counter = Counter(str(r["ProcessName"] or "(なし)") for r in group)
        header = " / ".join(f"{p}×{c}" for p, c in proc_counter.most_common())
        block = f"【プロセス内訳】{header}\n{summary}"
        for item in group:
            item["micro_10_summary"] = block


def build_micro_overview_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    groups: dict[tuple[str, int], list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        groups[(str(row["scenario"]), int(row["micro_rank"]))].append(row)
    out: list[dict[str, object]] = []
    for (scenario, rank), group in sorted(groups.items()):
        group.sort(key=lambda r: int(r["offset_in_micro"]))
        attack_n = sum(1 for r in group if r["label"] == 1)
        sample = group[0]
        out.append(
            {
                "scenario": scenario,
                "micro_rank": rank,
                "attack_in_micro": attack_n,
                "normal_in_micro": 10 - attack_n,
                "parent_session_id": sample["parent_session_id"],
                "minute_bucket": sample["minute_bucket"],
                "coarse_chunk_index": sample["coarse_chunk_index"],
                "micro_chunk_index": sample["micro_chunk_index"],
                "micro_chunk_key": sample["micro_chunk_key"],
                "main_processes": ", ".join(
                    p for p, _ in Counter(str(r["ProcessName"] or "(なし)") for r in group).most_common(5)
                ),
                "adoption_in_micro": ", ".join(
                    sorted({str(r["adoption"]) for r in group}, key=lambda x: (x == "attack", x))
                ),
                "micro_10_summary": sample["micro_10_summary"],
                "read_hint": (
                    "先に攻撃1件の位置を確認" if attack_n else "正常候補の塊として読む"
                ),
            }
        )
    return out


def build_summary_rows(master: list[dict[str, object]], lookup_notes: dict[str, str]) -> list[list[object]]:
    rows: list[list[object]] = [["区分", "件数", "備考"]]
    adoption_counts = Counter(str(r["adoption"]) for r in master)
    for key in sorted(adoption_counts, key=lambda name: (-adoption_counts[name], name)):
        rows.append([key, adoption_counts[key], ""])

    rows.append([])
    rows.append(["scenario", "生ログ参照", "件数"])
    for scenario in SCENARIOS:
        n = sum(1 for r in master if r["scenario"] == scenario.upper() and r.get("raw_json"))
        rows.append([scenario.upper(), lookup_notes.get(scenario, ""), n])

    rows.append([])
    rows.append(["scenario", "adoption", "count"])
    cross = Counter((r["scenario"], r["adoption"]) for r in master)
    for (scenario, adoption), count in sorted(cross.items()):
        rows.append([scenario, adoption, count])
    return rows


def autosize_columns(ws, max_width: int = 56, wide_cols: set[int] | None = None) -> None:
    wide_cols = wide_cols or set()
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        cap = 80 if col_idx in wide_cols else max_width
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), cap)


def style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def write_readable_master(ws, headers: list[str], dict_rows: list[dict[str, object]]) -> None:
    wide = set()
    for idx, h in enumerate(headers, start=1):
        if h in {"micro_10_summary", "raw_json", "template", "raw_one_liner"}:
            wide.add(idx)
    ws.append(headers)
    for record in dict_rows:
        ws.append([record.get(h, "") for h in headers])
    style_header_row(ws)
    for row_idx, record in enumerate(dict_rows, start=2):
        fill = fill_for_adoption(str(record.get("adoption", "")), str(record.get("color_hint", "")))
        if fill:
            for cell in ws[row_idx]:
                cell.fill = fill
        for col_idx in wide:
            ws.cell(row_idx, col_idx).alignment = WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws, wide_cols=wide)


def enrich_usecase_registry(master: list[dict[str, object]]) -> list[dict[str, object]]:
    registry = [dict(row) for row in USECASE_ROWS]
    for uc in registry:
        if uc["usecase_id"] == "補助":
            matches = [
                r
                for r in master
                if r["adoption"] in {"補助", "補助_周辺"}
                and r["ProcessName"] == uc["process"]
            ]
        else:
            matches = [r for r in master if r["adoption"] == uc["usecase_id"]]
        if matches:
            rep = matches[0]
            uc["representative_key"] = rep["event_key"]
            uc["representative_object_name"] = rep.get("ObjectName", "")
            uc["notes"] = f"{uc['notes']} / 400内 {len(matches)} 件"
    return registry


def load_session_indexes(jsonl_dir: Path | None) -> tuple[dict[str, dict[str, list[dict]]], dict[str, str]]:
    indexes: dict[str, dict[str, list[dict]]] = {}
    notes: dict[str, str] = {}
    for scenario in SCENARIOS:
        jsonl_path, gt_path, status = resolve_jsonl_path(scenario, ROOT, jsonl_dir)
        notes[scenario] = status
        if not jsonl_path:
            indexes[scenario] = {}
            continue
        print(f"Indexing {scenario}: {jsonl_path.name} ...")
        indexes[scenario] = build_session_event_index(jsonl_path, gt_path)
        print(f"  sessions={len(indexes[scenario])}")
    return indexes, notes


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build restoration phase 400-event Excel workbook.")
    parser.add_argument(
        "--jsonl-dir",
        type=Path,
        default=ROOT / "analysis_data" / "atlasv2_attack_runs" / "jsonl",
        help="Directory containing msft-security-h1-{s3,m4,m6,s4}.jsonl",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    jsonl_dir = args.jsonl_dir if args.jsonl_dir.exists() else None
    session_indexes, lookup_notes = load_session_indexes(jsonl_dir)

    master = load_master_rows(session_indexes, lookup_notes)
    if len(master) != 400:
        raise RuntimeError(f"expected 400 events, got {len(master)}")
    attach_micro_summaries(master)
    micro_rows = build_micro_overview_rows(master)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    ws_legend = wb.create_sheet("00_凡例")
    legend = [
        ["見方", "内容"],
        ["まず読む", "01b_micro10一覧 → 10件の塊"],
        ["次に読む", "01_400マスター → 1行=1ログ"],
        ["生ログ列", "ObjectName, raw_json 等（JSONL配置後に埋まる）"],
        [],
        ["色", "adoption", "意味"],
        ["緑", "U1-U3", "主採用"],
        ["黄", "U4, 補助", "境界・補助"],
        ["赤", "attack", "残attack2件"],
        ["灰", "除外", "背景"],
        [],
        ["JSONL配置先", str((ROOT / 'analysis_data/atlasv2_attack_runs/jsonl').relative_to(ROOT))],
    ]
    for scenario in SCENARIOS:
        legend.append([scenario.upper(), lookup_notes.get(scenario, "")])
    for row in legend:
        ws_legend.append(row)
    autosize_columns(ws_legend)

    ws_micro = wb.create_sheet("01b_micro10一覧")
    micro_headers = list(micro_rows[0].keys()) if micro_rows else []
    ws_micro.append(micro_headers)
    for record in micro_rows:
        ws_micro.append([record.get(h, "") for h in micro_headers])
    style_header_row(ws_micro)
    summary_col = micro_headers.index("micro_10_summary") + 1 if "micro_10_summary" in micro_headers else None
    for row_idx in range(2, 2 + len(micro_rows)):
        if summary_col:
            ws_micro.cell(row_idx, summary_col).alignment = WRAP
        if micro_rows[row_idx - 2].get("attack_in_micro"):
            for cell in ws_micro[row_idx]:
                cell.fill = FILL_ATTACK
    ws_micro.freeze_panes = "A2"
    autosize_columns(ws_micro, max_width=72, wide_cols={summary_col} if summary_col else set())

    ws_master = wb.create_sheet("01_400マスター")
    write_readable_master(ws_master, READABLE_HEADERS, master)

    ws_summary = wb.create_sheet("02_採用サマリ")
    for row in build_summary_rows(master, lookup_notes):
        ws_summary.append(row)
    style_header_row(ws_summary)
    autosize_columns(ws_summary)

    ws_uc = wb.create_sheet("03_U1-U4台帳")
    uc_headers = list(USECASE_ROWS[0].keys())
    uc_rows = enrich_usecase_registry(master)
    ws_uc.append(uc_headers)
    for record in uc_rows:
        ws_uc.append([record.get(h, "") for h in uc_headers])
    style_header_row(ws_uc)
    for row_idx in range(2, 2 + len(uc_rows)):
        uc_id = ws_uc.cell(row_idx, 1).value
        fill = FILL_CAUTION if uc_id == "U4" else FILL_UC if uc_id in {"U1", "U2", "U3"} else FILL_AUX
        for cell in ws_uc[row_idx]:
            cell.fill = fill
    ws_uc.freeze_panes = "A2"
    autosize_columns(ws_uc)

    ws_attack = wb.create_sheet("04_残attack2")
    attack_headers = list(ATTACK_ROWS[0].keys())
    ws_attack.append(attack_headers)
    for record in ATTACK_ROWS:
        enriched = dict(record)
        for row in master:
            if row["label"] == 1 and row["scenario"] == record["scenario"]:
                enriched["object_name"] = row.get("ObjectName", "")
                enriched["raw_one_liner"] = row.get("raw_one_liner", "")
                break
        ws_attack.append([enriched.get(h, "") for h in attack_headers])
    style_header_row(ws_attack)
    for row_idx in range(2, 2 + len(ATTACK_ROWS)):
        for cell in ws_attack[row_idx]:
            cell.fill = FILL_ATTACK
    ws_attack.freeze_panes = "A2"
    autosize_columns(ws_attack)

    ws_check = wb.create_sheet("05_復元チェックリスト")
    for row in CHECKLIST_ROWS:
        ws_check.append(row)
    style_header_row(ws_check)
    ws_check.freeze_panes = "A2"
    autosize_columns(ws_check)

    wb.save(OUT_XLSX)
    enriched = sum(1 for r in master if r.get("raw_json"))
    print(f"Wrote {OUT_XLSX} ({len(master)} events, raw_enriched={enriched})")
    if enriched == 0:
        print("NOTE: 攻撃シナリオ JSONL が無いため ObjectName/raw_json は空です。")
        print(f"      {ROOT / 'analysis_data/atlasv2_attack_runs/jsonl'} に msft-security-h1-*.jsonl を置いて再実行してください。")


if __name__ == "__main__":
    main()
